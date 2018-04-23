import openpyxl
import time
# import json
from datetime import datetime
from functools import reduce
from operator import itemgetter

import tableextraction
import exceladapter
import translators
import excelbuilder
import datastructures
import dateintelligence

# abrir archivos
NUEVO = "nuevo"
HISTORICO = "histórico"

current_date = datetime.now()
first_day_of_current_month = datetime(current_date.year, current_date.month, 1)
# last_date_of_month = datetime(current_date.year,current_date.month,1) + relativedelta(months=1,days=-1)
# TODO: Generalizar
last_date_of_month = datetime(2018, 4, 30)

def get_version_from_code(raw_code):
    if "de" in raw_code.split("-")[3]:
        return HISTORICO
    return NUEVO


def configure_sheet(sheet_builder):
    sheet_builder.configure_column("A", "Ciudad", 'city')
    sheet_builder.configure_column("B", "Cliente", 'customer')
    sheet_builder.configure_column("C", "Dirección", 'address')
    sheet_builder.configure_column("D", "Teléfono", 'phone')

    sheet_builder.configure_column("E", "Fecha de Compra", 'date_of_purchase')
    sheet_builder.configure_column("F", "Fecha de Vencimiento", 'due_date')
    sheet_builder.configure_column("G", "Valor de compra", 'total_purchase_value')

    sheet_builder.configure_column("H", "Orden de Compra", 'order')
    sheet_builder.configure_column("I", "Última Cobranza", 'last_collection')

    sheet_builder.configure_column("J", "Cuotas", 'plan')
    sheet_builder.configure_column("K", "Saldo Total", 'debt_balance')
    sheet_builder.configure_column("L", "Cuota a pagar", 'current_payment')
    sheet_builder.configure_column("M", "Valor de cuota", 'payment')
    sheet_builder.configure_column("N", "Saldo vencido", 'past_due_debt')  # revisar
    sheet_builder.configure_column("O", "Deuda impaga a la fecha", 'overdue_balance')

    sheet_builder.configure_column("P", "Monto total a cobrar", 'amount_to_collect')

calendar_ops = dateintelligence.CalendarOperations()

errors = []
customers = {}
collections = datastructures.HashOfLists()
bills = {}

accounts_to_collect = {
    "C": [],
    "D": [],
    "I": []
}

customers_reader = exceladapter.excelreader.ExcelReader('inputs/Clientes.xlsx')
collections_reader = exceladapter.excelreader.ExcelReader('inputs/Cobranza.xlsx')
pending_bills_reader = exceladapter.excelreader.ExcelReader('inputs/Factura.xlsx')
accounts_to_collect_reader = exceladapter.excelreader.ExcelReader('inputs/Cuentas a Cobrar.xlsx')

customers_sheet = customers_reader.get_sheet('Sheet0')
collections_sheet = collections_reader.get_sheet('hoja1')
pending_bills_sheet = pending_bills_reader.get_sheet('hoja1')
accounts_to_collect_sheet = accounts_to_collect_reader.get_sheet('hoja1')

customers_unpacker = tableextraction.TableUnpacker(customers_sheet)

for row_unpacker in customers_unpacker.read_rows(2):
    customer = {}

    name = row_unpacker.get_value_at(1)

    customer['address'] = row_unpacker.get_value_at(8)
    customer['city'] = row_unpacker.get_value_at(28)
    customer['phone'] = row_unpacker.get_value_at(12)

    customers[name] = customer

collections_unpacker = tableextraction.TableUnpacker(collections_sheet)

for row_unpacker in collections_unpacker.read_rows(2):
    collection = {}

    document = row_unpacker.get_value_at(2)
    raw_code = row_unpacker.get_value_at(5)

    sales_order = ""

    if "-" in raw_code:
        version = NUEVO
        order_and_payments = raw_code.split("-")
        sales_order = order_and_payments.pop(0)
        payments = order_and_payments
    else:
        payments = ""
        version = HISTORICO

    if sales_order and len(sales_order) != 5:
        error = "Cobranza con orden de compra errónea (%s). Documento: %s" % (sales_order, document)
        errors.append(error)

    # Fecha	Comprobante	Cliente	Total	Observaciones
    collection['version'] = version
    collection['date'] = row_unpacker.get_value_at(1)
    collection['customer'] = row_unpacker.get_value_at(3)
    collection['amount'] = row_unpacker.get_value_at(4)
    collection['sales_order'] = sales_order
    collection['payments'] = payments

    collections.append(sales_order, collection)

bills_unpacker = tableextraction.TableUnpacker(pending_bills_sheet)

for row_unpacker in bills_unpacker.read_rows(2):

    document = row_unpacker.get_value_at(4)
    raw_code = row_unpacker.get_value_at(11)

    if not raw_code:
        errors.append("Factura sin descripción. Documento: %s" % (document))
        continue

    if "de" in raw_code.split("-")[3]:
        version = HISTORICO
    else:
        version = NUEVO

    if version == HISTORICO:
        continue

    sales_order = raw_code.split("-")[2]

    if not sales_order:
        errors.append("Factura sin orden de compra. Documento: %s" % (document))
        continue

    if len(sales_order) != 5:
        error = "Factura con orden de compra errónea (%s). Documento: %s" % (sales_order, document)
        errors.append(error)

    amount = row_unpacker.get_value_at(18)

    if sales_order in bills and version == NUEVO:
        errors.append("Orden de compra repetida. Documento: %s. Orden de compra: %s" % (document, sales_order))
        continue

    bills[sales_order] = amount

accounts_to_collect_unpacker = tableextraction.TableUnpacker(accounts_to_collect_sheet)

for row_unpacker in accounts_to_collect_unpacker.read_rows(2):

    account_to_collect = {}

    document_date = row_unpacker.get_value_at(1)
    due_date = row_unpacker.get_value_at(2)
    document = row_unpacker.get_value_at(3)
    customer = row_unpacker.get_value_at(4)

    customer_data = customers[customer]

    city = customer_data['city']
    phone = customer_data['phone']
    address = customer_data['address']

    if "Cobranza" in document:
        continue

    raw_code = row_unpacker.get_value_at(9)

    if not raw_code:
        # ya lo toma de las facturas, no debería en un listado sin que esté en el otro
        # errors.append("Cuenta a cobrar sin descripción. Documento: %s" % (document))
        continue

    version = get_version_from_code(raw_code)

    if version == HISTORICO and due_date > last_date_of_month:
        continue

    list_of_codes = raw_code.split("-")

    collection_account = list_of_codes[0]
    collection_person = list_of_codes[1]
    sales_order = list_of_codes[2]

    collections_for_order = []
    last_collection_date = "Sin cobranza previa"
    current_payment_number = 1

    if sales_order and sales_order in collections:
        collections_for_order = collections[sales_order]

    if sales_order and len(collections_for_order) > 0:
        last_collection_date = sorted(collections_for_order, key=lambda x: x['date'], reverse=True)[0]['date']

        if version == NUEVO:
            previous_payments = reduce(lambda x, y: x + y, list(map(lambda x: x['payments'], collections_for_order)),
                                       [])
            previous_payments_without_advances_str = filter(lambda x: x != 'E', previous_payments)
            previous_payments_without_advances = [int(x) for x in previous_payments_without_advances_str]
            current_payment_number = max(previous_payments_without_advances, default=0) + 1

    if version == HISTORICO:
        current_payment_number, plan = list_of_codes[3].split(" de ")
        current_payment_number = int(current_payment_number)
        plan = int(plan)
        payment_amount = float(row_unpacker.get_value_at(8))
        total_purchase_value = payment_amount * int(plan)

        debt_balance = ""
        advance_payment = ""

    else:
        if len(list_of_codes) < 5:
            error = "Cuenta a cobrar sin valor de cuota. Documento: %s - Descripción: %s" % (document, raw_code)
            errors.append(error)
            continue

        plan = int(list_of_codes[3])
        payment_amount = float(list_of_codes[4])
        debt_balance = row_unpacker.get_value_at(8)
        total_purchase_value = bills[sales_order]
        paid_amount = total_purchase_value - debt_balance
        advance_payment = total_purchase_value - (plan * payment_amount)

        if (advance_payment < 0):
            error = "El monto de venta es menor al plan de pagos. Documento: %s - Valor: %s. Plan: %s - Cuota: %s. Diferencia: %s" % (
            document, total_purchase_value, plan, payment_amount, advance_payment)
            errors.append(error)

    # TODO: months_from_sale

    last_due_date = calendar_ops.add_months(due_date, plan)

    overdue_balance = 0
    past_due_debt = ""

    if version == NUEVO:
        due_dates = calendar_ops.list_of_due_date(due_date, plan)
        #  i = next(i + 1 for i,v in enumerate(l) if v > 0)
        due_payments = next((plan - i for i, v in enumerate(reversed(due_dates)) if v < first_day_of_current_month), 0)
        past_due_debt = advance_payment + (due_payments * payment_amount)
        overdue_balance = past_due_debt - paid_amount

        '''
        if customer == "HUENTEN JORGE":
            print('due payments', due_payments, "plan", plan, "due_dates", due_dates)
            print('adelanto', advance_payment, "cuota", payment_amount)
            print('paid amount', paid_amount, 'past due debt', past_due_debt, 'overdue', overdue_balance)
        '''

    if isinstance(last_collection_date, datetime):
        last_collection_date = last_collection_date.strftime("%d/%m/%Y")

    account_to_collect['version'] = version

    account_to_collect['document'] = document
    account_to_collect['date_of_purchase'] = document_date.strftime("%d/%m/%Y")

    account_to_collect['due_date_datetime'] = due_date
    account_to_collect['due_date'] = due_date.strftime("%d/%m/%Y")

    account_to_collect['customer'] = customer

    account_to_collect['city'] = city
    account_to_collect['address'] = address
    account_to_collect['phone'] = phone

    account_to_collect['account'] = collection_account
    account_to_collect['person'] = collection_person

    '''
    if collection_account == "D":
        print(collection_account, collection_person)
    '''

    account_to_collect['order'] = sales_order

    account_to_collect['last_collection'] = last_collection_date
    account_to_collect['total_purchase_value'] = total_purchase_value
    account_to_collect['plan'] = plan
    account_to_collect['advance_payment'] = advance_payment

    account_to_collect['debt_balance'] = debt_balance

    account_to_collect['current_payment'] = current_payment_number
    account_to_collect['payment'] = payment_amount
    account_to_collect['past_due_debt'] = past_due_debt
    account_to_collect['overdue_balance'] = overdue_balance

    account_to_collect['amount_to_collect'] = float(payment_amount) + float(overdue_balance)

    if not account_to_collect['city'] or not account_to_collect['customer']:
        print("NOT CITY OR CUSTOMER", account_to_collect['city'], account_to_collect['customer'])

    accounts_to_collect[collection_account].append(account_to_collect)

for error in errors:
    print("WARNING:", error)

# crear excel de cobranzas
collections_filename = 'outputs/cuentas_a_cobrar_' + time.strftime("%Y%m%d-%H%M%S") + '.xlsx'
collections_excelwriter = exceladapter.ExcelWriter(collections_filename)

generated_sheet_C = collections_excelwriter.create_sheet('Créditos')
generated_sheet_DH = collections_excelwriter.create_sheet('Débitos Horacio')
generated_sheet_DF = collections_excelwriter.create_sheet('Débitos Facundo')
generated_sheet_I = collections_excelwriter.create_sheet('ICBC')

# print(accounts_to_collect['C'])

sorted_accounts_C = sorted(accounts_to_collect['C'],
                           key=lambda x: (x['city'], x['customer'], x['order'], x['due_date_datetime']))

collections_builder_C = excelbuilder.BasicBuilder(generated_sheet_C, sorted_accounts_C)
configure_sheet(collections_builder_C)
collections_builder_C.build()

sorted_accounts_D = sorted(accounts_to_collect['D'],
                           key=lambda x: (x['city'], x['customer'], x['order'], x['due_date_datetime']))

sorted_accounts_D_H = filter(lambda x: x['person'] == 'H', sorted_accounts_D)
sorted_accounts_D_F = filter(lambda x: x['person'] == 'F', sorted_accounts_D)

collections_builder_DH = excelbuilder.BasicBuilder(generated_sheet_DH, sorted_accounts_D_H)
configure_sheet(collections_builder_DH)
collections_builder_DH.build()

collections_builder_DF = excelbuilder.BasicBuilder(generated_sheet_DF, sorted_accounts_D_F)
configure_sheet(collections_builder_DF)
collections_builder_DF.build()


sorted_accounts_I = sorted(accounts_to_collect['I'],
                           key=lambda x: (x['city'], x['customer'], x['order'], x['due_date_datetime']))

collections_builder_I = excelbuilder.BasicBuilder(generated_sheet_I, sorted_accounts_I)
configure_sheet(collections_builder_I)
collections_builder_I.build()

collections_excelwriter.save()

errors_filename = 'outputs/errors_' + time.strftime("%Y%m%d-%H%M%S") + '.txt'
with open(errors_filename, 'w') as f:
    for error in errors:
        f.write(error)
        f.write("\n")
